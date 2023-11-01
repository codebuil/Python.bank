
import tkinter as tk
from tkinter import messagebox
import openpyxl
import os

def save_to_excel(data, file_name):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Account"

    headers = ['Entity', 'Date', 'Reference', 'Value']
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=header)

    for row, row_data in enumerate(data, start=2):
        for col, value in enumerate(row_data, start=1):
            sheet.cell(row=row, column=col, value=value)

    wb.save(file_name)
    messagebox.showinfo("Success", f"Data saved to '{file_name}'")

def add_data_to_memory(entity, date, reference, value, file_name):
    data = []
    if os.path.exists(file_name):
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active

        for row in sheet.iter_rows(values_only=True):
            data.append([row[0], row[1], row[2], row[3]])

    data.append([entity, date, reference, value])
    save_to_excel(data, file_name)

def display_entity_data(entity, file_name):
    data = []
    if os.path.exists(file_name):
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active

        for row in sheet.iter_rows(values_only=True):
            if row[0] == entity:
                data.append(list(row))

    entity_data = ""
    total_sum = 0
    entity_data += f"Entity {entity} data:\n"
    entity_data += f"{'Entity':10}{'Date':15}{'Reference':15}{'Value':15}{'Total:':15}\n"
    for row in data:
        if row[0] == entity:
            total_sum += row[3]
            entity_data += f"{row[0]:10}{row[1]:15}{row[2]:15}{row[3]:15}{total_sum:15}\n"

    entity_data += f"{'SUM':40}{total_sum}"
    messagebox.showinfo("Entity Data", entity_data)

def delete_data(entity, file_name, delete_index):
    data = []
    if os.path.exists(file_name):
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active

        for row in sheet.iter_rows(values_only=True):
            if row[0] == entity:
                data.append(list(row))

        data2 = [row for row in sheet.iter_rows(values_only=True) if row[0] != entity]

        if 0 <= delete_index < len(data):
            data.pop(delete_index)

        save_to_excel(data + data2, file_name)

def add_entity():
    def add_entity_data():
        entity = entity_entry.get()
        date = date_entry.get()
        reference = reference_entry.get()
        value = float(value_entry.get())

        add_data_to_memory(entity, date, reference, value, "account.xlsx")
        add_entity_window.destroy()

    add_entity_window = tk.Toplevel(root)
    add_entity_window.title("Add Entity Data")

    entity_label = tk.Label(add_entity_window, text="Entity:")
    entity_label.grid(row=0, column=0)
    entity_entry = tk.Entry(add_entity_window)
    entity_entry.grid(row=0, column=1)

    date_label = tk.Label(add_entity_window, text="Date:")
    date_label.grid(row=1, column=0)
    date_entry = tk.Entry(add_entity_window)
    date_entry.grid(row=1, column=1)

    reference_label = tk.Label(add_entity_window, text="Reference:")
    reference_label.grid(row=2, column=0)
    reference_entry = tk.Entry(add_entity_window)
    reference_entry.grid(row=2, column=1)

    value_label = tk.Label(add_entity_window, text="Value:")
    value_label.grid(row=3, column=0)
    value_entry = tk.Entry(add_entity_window)
    value_entry.grid(row=3, column=1)

    submit_button = tk.Button(add_entity_window, text="Submit", command=add_entity_data)
    submit_button.grid(row=4, columnspan=2)

def display_entity():
    def show_entity_data():
        entity = display_entry.get()
        display_entity_data(entity, "account.xlsx")
        display_entity_window.destroy()

    display_entity_window = tk.Toplevel(root)
    display_entity_window.title("Display Entity Data")

    display_label = tk.Label(display_entity_window, text="Entity to display:")
    display_label.pack()
    display_entry = tk.Entry(display_entity_window)
    display_entry.pack()

    display_button = tk.Button(display_entity_window, text="Display", command=show_entity_data)
    display_button.pack()

def delete_entity():
    def delete_entity_data():
        entity = delete_entry.get()
        delete_index = int(delete_index_entry.get())

        delete_data(entity, "account.xlsx", delete_index)
        delete_entity_window.destroy()

    delete_entity_window = tk.Toplevel(root)
    delete_entity_window.title("Delete Entity Data")

    delete_label = tk.Label(delete_entity_window, text="Entity to delete:")
    delete_label.grid(row=0, column=0)
    delete_entry = tk.Entry(delete_entity_window)
    delete_entry.grid(row=0, column=1)

    delete_index_label = tk.Label(delete_entity_window, text="Index to delete:")
    delete_index_label.grid(row=1, column=0)
    delete_index_entry = tk.Entry(delete_entity_window)
    delete_index_entry.grid(row=1, column=1)

    delete_button = tk.Button(delete_entity_window, text="Delete", command=delete_entity_data)
    delete_button.grid(row=2, columnspan=2)

root = tk.Tk()
root.title("Entity Data Management")
root.configure(bg="brown", width=800, height=600)  # Define a cor de 

add_button = tk.Button(root, text="Add Entity", command=add_entity)
add_button.pack()

display_button = tk.Button(root, text="List Entity Data", command=display_entity)
display_button.pack()

delete_button = tk.Button(root, text="Delete Entity Data", command=delete_entity)
delete_button.pack()

exit_button = tk.Button(root, text="Exit", command=root.quit)
exit_button.pack()

root.mainloop()
