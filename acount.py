import openpyxl
import os
from openpyxl.utils import get_column_letter

def save_to_excel(data, file_name):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Account"

    headers = ['Entity', 'Date', 'Reference', 'Value']
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=header)

    for row, row_data in enumerate(data, start=2):
        for col, value in enumerate(row_data, start=1):
            sheet.cell(row=row, column=col, value=value)

    wb.save(file_name)
    print(f"Data saved to '{file_name}'")

def add_data_to_memory(datas, entity, file_name):
    data = []
    if os.path.exists(file_name):
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active

    
  
        for row in sheet.iter_rows(values_only=True):
            data.append([row[0], row[1], row[2], row[3]])
  
    while True:
        user_input = input("Enter data (date, reference, value) for the entity (press Enter to stop): ")
        if user_input:
            date, reference, value = user_input.split(',')
            data.append([entity, date.strip(), reference.strip(), float(value.strip())])
        else:
            save_to_excel(data, file_name)
            break

def display_entity_data(file_name,data, entity):
    data = []
    if os.path.exists(file_name):
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active

    
        for row in sheet.iter_rows(values_only=True):
            if row[0]==entity:
                 data.append(list(row))


    total_sum = 0
    print(f"Entity {entity} data:")
    print(f"{'Entity':10}{'Date':15}{'Reference':15}{'Value':15}{'Total:':15}")
    for row in data:
        if row[0] == entity:
            total_sum += row[3]
            print(f"{row[0]:10}{row[1]:15}{row[2]:15}{row[3]:15}{total_sum:15}")
            

    print(f"{'SUM':40}{total_sum}")

def delete_data(data, entity, file_name):
    data = []
    data2 = []
    deletedes= []
    if os.path.exists(file_name):
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active

    
    
        for row in sheet.iter_rows(values_only=True):
            if row[0]==entity:
                data.append(list(row))
            else:
                data2.append(list(row))

    print(f"Entries for Entity {entity}:")
    print(f"{'No.':6}{'Entity':10}{'Date':15}{'Reference':15}{'Value':15}")
    
    for i, row in enumerate(data, start=0):
        print(f"{i:5}={row[0]:10}{row[1]:15}{row[2]:15}{row[3]:15}")
        

    while True:
        delete_index = input("Enter the line number to delete (press Enter to stop): ")
        if delete_index:
            if 0 <= int(delete_index) < len(data):
  
                deletedes.append(int(delete_index))
       
            
        else:
            data3= []
            print("\x1bc\x1b[43;30m")
            print(f"{'Entity':10}{'Date':15}{'Reference':15}{'Value':15}")
            for i,row in enumerate(data,start=0):
                t=0
                for n in deletedes:
                    if i==n:
                        t=1
                if t==0:    
                    data3.append([row[0], row[1], row[2], row[3]])
            if len(data2)>0:
                for row in enumerate(data2):
                    try:
                        data3.append([row[0], row[1], row[2], row[3]])
                    except:
                        pass
            save_to_excel(data3, file_name)
            break

if 0==0:
    print("\x1bc\x1b[43;30m")

    data = []

    while True:
        print("MENU:")
        print("1. Add Entity")
        print("2. List Entity Data")
        print("3. Delete Entity Data")
        print("4. Exit")
        choice = input("Enter your choice: ")

        if choice == '1':
            entity = input("Enter the entity: ")
            add_data_to_memory(data, entity, "account.xlsx")
            data = []
        elif choice == '2':
            entity = input("Enter the entity to display data: ")
            display_entity_data("account.xlsx",data, entity)
        elif choice == '3':
            entity = input("Enter the entity to delete: ")
            delete_data(data, entity, "account.xlsx")
        elif choice == '4':
            break
        else:
            print("Invalid choice. Please enter a valid option.")

